import re
import openpyxl

# 과제 (필수)
# 01_data/train.xlsx 엑셀 파일을 읽어서, 다음과 같이 이름에 들어
# 있는 정보를 기반으로 4개의 쉬트를 만들어 각 행의 정보를 복사하세요.
# Mr (쉬트명은 남성 : Mr.)  Miss (쉬트명은 미혼여성 : Miss)
# Mrs (쉬트명은 기혼여성 : Mrs)  Others (쉬트명은 기타 : 없음 )
work_book = openpyxl.load_workbook('./regex/train.xlsx')
work_sheet = work_book.active

# 파일 작성
# 4개의 시트 작성
wb = openpyxl.Workbook()
work_sheet_man = wb.active
work_sheet_man.column_dimensions['D'].width = 70
work_sheet_man.title = '남성'
work_sheet_solo_women = wb.create_sheet(title='미혼여성')
work_sheet_solo_women.column_dimensions['D'].width = 70
work_sheet_married_women = wb.create_sheet(title='기혼여성')
work_sheet_married_women.column_dimensions['D'].width = 70
work_sheet_others = wb.create_sheet(title='기타')
work_sheet_others.column_dimensions['D'].width = 70

pattern = re.compile(r' [A-Za-z]+\.')

# 남자 총인원수 survied
man_survived, man_unsurvived = 0, 0
single_survived, single_unsurvived = 0, 0
married_survived, married_unsurvived = 0, 0
others_survived, others_unsurvived = 0, 0

for each_row in work_sheet.rows:
    # print(each_row[3].value)

    data = pattern.findall(each_row[3].value)
    # print(data)  컬럼에서 정규식과 매치된 부분 출력됨

    # train 쉬트의 첫 행은 항목 이름이 있으므로, 이를 각 쉬트에 복사하기
    if each_row[0].row == 1:
        work_sheet_man.append([each_item.value for each_item in each_row])
        work_sheet_solo_women.append(
            [each_item.value for each_item in each_row])
        work_sheet_married_women.append(
            [each_item.value for each_item in each_row])
        work_sheet_others.append([each_item.value for each_item in each_row])
    else:
        if len(data) > 0:    # 이 부분 반드시 필요
            if data[0] == ' Mr.':
                work_sheet_man.append(
                    [each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    man_survived += 1
                else:
                    man_unsurvived += 1
            elif data[0] == ' Miss.':
                work_sheet_solo_women.append(
                    [each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    single_survived += 1
                else:
                    single_unsurvived += 1
            elif data[0] == ' Mrs.':
                work_sheet_married_women.append(
                    [each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    married_survived += 1
                else:
                    married_unsurvived += 1
            else:
                work_sheet_others.append(
                    [each_item.value for each_item in each_row])
                if each_row[1].value == 1:
                    others_survived += 1
                else:
                    others_unsurvived += 1


# 도전 과제 (옵션)
# 이외에 추가로 한 개 쉬트를 더 만들어서, (쉬트명은 보고서) 다음 정보를 기록하세요.
work_sheet_report = wb.create_sheet(title='보고서')

# 분류	생존자수	사망자수	생존률
work_sheet_report.append(['분류', '생존자수', '사망자수', '생존률'])
# 남성	#	#	#.## %
survived_rate = "%.2f%%" % (
    man_survived / (man_survived + man_unsurvived) * 100)
print("%s %d %d %s" % ('남성', man_survived, man_unsurvived, survived_rate))
work_sheet_report.append(['남성', man_survived, man_unsurvived, survived_rate])
# 미혼여성	#	#	#.## %
single_survived_rate = "%.2f%%" % (single_survived / (single_survived + single_unsurvived) * 100)
print("%s %d %d %s" % ('미혼여성', single_survived,single_unsurvived, single_survived_rate))
work_sheet_report.append(['미혼여성', single_survived, single_unsurvived, single_survived_rate])
# 기혼여성	#	#	#.## %
married_survived_rate = "%.2f%%" % (married_survived / (married_survived + married_unsurvived) * 100)
print("%s %d %d %s" % ('기혼여성', married_survived, married_unsurvived, married_survived_rate))
work_sheet_report.append(['기혼여성', married_survived, married_unsurvived, married_survived_rate])
# 기타	#	#	#.## %
others_survived_rate = "%.2f%%" % (others_survived / (others_survived + others_unsurvived) * 100)
print("%s %d %d %s" % ('기타', others_survived,others_unsurvived, others_survived_rate))
work_sheet_report.append(['기타', others_survived, others_unsurvived, others_survived_rate])

# - 생존률은 사망자 수 / 생존자 수 + 사망자 수
# - #.## % 는 소숫점 두자리수까지 표기하라는 의미로 예를 들어 32.12 % 등

wb.save(filename='./regex/train_gender.xlsx')
wb.close()
work_book.close()
