from openpyxl import Workbook

# openpyxl WorkBook 객체 생성
excel_file = Workbook()
print(excel_file.sheetnames)

# 기존 시트 하나 없애기
excel_file.remove(excel_file['Sheet'])
# excel_file.remove(excel_file[excel_file.sheetnames]) 로는 지울 수 없음

# 첫번째 시트 생성하기
sheet1 = excel_file.create_sheet(index=0, title="Column")
# 새로운 시트 생성하기
sheet2 = excel_file.create_sheet(index=1, title="매출표")
print(excel_file.sheetnames)

excel_file.save("./resources/test2.xlsx")
