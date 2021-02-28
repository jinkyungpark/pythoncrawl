from openpyxl import Workbook
import openpyxl.utils.cell as utils
from openpyxl.styles import Alignment, Font

# openpyxl WorkBook 객체 생성
excel_file = Workbook()
print(excel_file.sheetnames)

# 기존 시트 하나 없애기
excel_file.remove(excel_file['Sheet'])
# excel_file.remove(excel_file[excel_file.sheetnames]) 로는 지울 수 없음

# 첫번째 시트 생성하기
sheet1 = excel_file.create_sheet(index=0, title="Column")
# 새로운 시트 생성하기
sheet2 = excel_file.create_sheet(index=1, title="매출표")
print(excel_file.sheetnames)


# 셀 서식 지정 3행 6열
for col in sheet1.iter_cols(min_col=1, max_col=6, min_row=1, max_row=3):
    for each_cell in col:
        # 셀에 들어갈 값 지정 (컬럼명 가져와서 삽입)
        each_cell.value = utils.get_column_letter(each_cell.column)
        # 정렬 기준 설정
        each_cell.alignment = Alignment(horizontal='right', vertical='center')
        # 폰트 설정
        each_cell.font = Font(bold=True, name="Arial",
                              size=12, underline='single', color='1bb638')

# 셀 서식 지정 6행 3열
for row in sheet2.iter_cols(min_col=1, max_col=3, min_row=1, max_row=6):
    for each_cell in row:
        # 셀에 들어갈 값 지정 (컬럼명 가져와서 삽입)
        each_cell.value = each_cell.row
        # 정렬 기준 설정
        each_cell.alignment = Alignment(horizontal='center', vertical='center')
        # 폰트 설정
        each_cell.font = Font(italic=True, name="Consoras",
                              size=10, color='ff0000')

excel_file.save("./resources/test2.xlsx")
